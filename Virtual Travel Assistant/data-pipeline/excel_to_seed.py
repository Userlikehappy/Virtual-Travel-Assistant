"""
Convert Excel data to ai-engine/data/seed/{locations,food}.json
Reads the official tourism Excel file directly — no web crawl needed.

Usage:
    python data-pipeline/excel_to_seed.py
"""

import json
import math
import sys
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

EXCEL_PATH = Path(__file__).parent.parent / "thong_tin_danh_thang_dia_diem_du_lich_tai_thanh_ph.xlsx"
OUT_DIR    = Path(__file__).parent.parent / "ai-engine" / "data" / "seed"

# === Geographic bounds — only keep locations within these areas ===
AREAS = {
    "Đà Nẵng":  {"lat": 16.047, "lng": 108.206, "radius_km": 60},
    "Hội An":   {"lat": 15.880, "lng": 108.338, "radius_km": 30},
    "Huế":      {"lat": 16.464, "lng": 107.591, "radius_km": 40},
}

def haversine(lat1, lng1, lat2, lng2) -> float:
    R = 6371.0
    φ1, φ2 = math.radians(lat1), math.radians(lat2)
    dφ = math.radians(lat2 - lat1)
    dλ = math.radians(lng2 - lng1)
    a = math.sin(dφ/2)**2 + math.cos(φ1)*math.cos(φ2)*math.sin(dλ/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))

def nearest_city(lat, lng):
    best, best_dist = None, 9999
    for city, cfg in AREAS.items():
        d = haversine(lat, lng, cfg["lat"], cfg["lng"])
        if d < best_dist:
            best_dist, best = d, city
    # Only assign city if within radius
    if best and best_dist <= AREAS[best]["radius_km"]:
        return best
    return None

def parse_coord(val) -> float:
    """Parse a coordinate cell value that might be str/float/int with dots as separators."""
    if val is None:
        return 0.0
    s = str(val).strip().replace(" ", "")
    # Remove trailing/leading junk
    s = re.sub(r"[^\d.\-]", "", s)
    if not s:
        return 0.0
    try:
        f = float(s)
        return f
    except ValueError:
        return 0.0

def parse_price(val) -> float:
    """Extract a numeric price from messy strings like '150.000–300.000', '200000.0'."""
    if val is None:
        return 0.0
    s = str(val).strip()
    if "miễn phí" in s.lower() or "free" in s.lower():
        return 0.0
    # Find all numbers (possibly with dots as thousand separators)
    nums = re.findall(r"[\d]+(?:[.,][\d]+)*", s)
    values = []
    for n in nums:
        try:
            # Remove dots/commas used as thousand separators (if >= 1000 range)
            cleaned = n.replace(".", "").replace(",", "")
            v = float(cleaned)
            if v > 10:  # ignore single digits
                values.append(v)
        except ValueError:
            pass
    if not values:
        return 0.0
    return max(values)  # return the maximum (upper bound of range)

def parse_bool(val, true_keywords=("có", "yes", "true", "1")) -> bool:
    if val is None:
        return False
    s = str(val).strip().lower()
    return any(k in s for k in true_keywords)

CATEGORY_MAP = {
    "văn hóa": "culture",
    "làng nghề": "culture",
    "làng truyền thống": "culture",
    "di tích lịch sử": "heritage",
    "di sản": "heritage",
    "kiến trúc": "heritage",
    "lịch sử": "heritage",
    "di tích": "heritage",
    "chùa": "spiritual",
    "đình": "spiritual",
    "đền": "spiritual",
    "tâm linh": "spiritual",
    "thờ": "spiritual",
    "bảo tàng": "museum",
    "triển lãm": "museum",
    "biển": "beach",
    "bãi tắm": "beach",
    "bãi biển": "beach",
    "núi": "nature",
    "rừng": "nature",
    "thiên nhiên": "nature",
    "sinh thái": "nature",
    "giải trí": "entertainment",
    "vui chơi": "entertainment",
    "café": "cafe",
    "cafe": "cafe",
    "cà phê": "cafe",
    "ẩm thực": "food",
    "nhà hàng": "food",
    "quán ăn": "food",
    "thể thao": "sports",
    "mua sắm": "shopping",
    "chợ": "shopping",
}

def classify_category(loai: str) -> str:
    if not loai:
        return "sightseeing"
    loai_lower = loai.lower()
    for keyword, cat in CATEGORY_MAP.items():
        if keyword in loai_lower:
            return cat
    return "sightseeing"

def classify_environment(loai: str, rain_ok: str) -> str:
    loai_lower = (loai or "").lower()
    rain_lower = (rain_ok or "").lower()
    if any(k in loai_lower for k in ["bảo tàng", "cafe", "nhà hàng", "quán ăn", "trong nhà"]):
        return "Indoor"
    if "có thể" in rain_lower and "trong nhà" in rain_lower:
        return "Indoor"
    if "hạn chế" in rain_lower or "không nên" in rain_lower:
        return "Outdoor"
    if any(k in loai_lower for k in ["biển", "bãi tắm", "núi", "rừng", "sinh thái"]):
        return "Outdoor"
    return "Outdoor"

def parse_hours(val) -> str:
    if not val:
        return ""
    s = str(val).strip()
    # Clean up formula artifacts
    if s.startswith("="):
        return ""
    # Standardize separators
    s = s.replace("–", "-").replace("—", "-")
    # Extract HH:MM patterns
    times = re.findall(r"\d{1,2}:\d{2}", s)
    if len(times) >= 2:
        return f"{times[0]}-{times[1]}"
    if "tự do" in s.lower() or "cả ngày" in s.lower() or "24" in s:
        return "00:00-23:59"
    return s[:30] if s else ""


def read_locations_sheet(wb) -> list:
    """Read 'Địa điểm ' sheet — col layout: 2=name, 3=loai, 4=desc, 7=district, 8=city, 9=lat, 10=lng, 12=hours, 15=ticket, 18=rating, 20=rain, 25=car_parking"""
    ws = wb["Địa điểm "]
    records = []
    seen = set()

    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, 2).value or "").strip()
        if not name or name in seen:
            continue

        loai     = str(ws.cell(r, 3).value or "").strip()
        desc     = str(ws.cell(r, 4).value or "").strip()[:500]
        district = str(ws.cell(r, 7).value or "").strip().rstrip(",")
        city_raw = str(ws.cell(r, 8).value or "").strip()
        lat_raw  = ws.cell(r, 9).value   # labeled "Kinh độ" but actually lat
        lng_raw  = ws.cell(r, 10).value  # labeled "Vĩ độ" but actually lng
        hours    = parse_hours(ws.cell(r, 12).value)
        ticket   = parse_price(ws.cell(r, 15).value)
        rating   = parse_coord(ws.cell(r, 18).value)
        rain_ok  = str(ws.cell(r, 20).value or "").strip()
        car_park = parse_bool(ws.cell(r, 25).value)

        lat = parse_coord(lat_raw)
        lng = parse_coord(lng_raw)

        # Swap if clearly reversed
        if lat > 100 and lng < 30:
            lat, lng = lng, lat

        # Validate Vietnam coordinate range
        if not (8 <= lat <= 24 and 102 <= lng <= 110):
            continue

        city = nearest_city(lat, lng)
        if not city:
            continue  # outside our target areas

        seen.add(name)
        category = classify_category(loai)
        environment = classify_environment(loai, rain_ok)

        record = {
            "name": name,
            "category": category,
            "type": "location",
            "environment": environment,
            "district": district or city_raw,
            "city": city,
            "lat": round(lat, 6),
            "lng": round(lng, 6),
            "description": desc,
            "operating_hours": hours,
            "estimated_cost": ticket,
            "estimated_rating": round(rating, 1) if rating > 0 else 0,
            "car_parking": car_park,
            "motorbike_parking": True,
            "meal_time": [],
            "season_tags": ["all_year"],
            "style": "heritage" if category in ("heritage", "spiritual", "culture") else "trending",
        }
        records.append(record)

    return records


def read_food_sheet(wb) -> list:
    """Read 'Ăn uống' sheet — data starts row 1, col structure varies slightly.
    Strategy: find name (first non-empty short text col), find lat/lng by value range."""
    ws = wb["Ăn uống"]
    records = []
    seen = set()

    for r in range(1, ws.max_row + 1):
        # Collect all cell values
        cells = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        str_cells = [str(v or "").strip() for v in cells]

        # Find name: first non-empty string that is not a category keyword or number
        name = ""
        name_col = -1
        for i, s in enumerate(str_cells[:6]):
            if s and len(s) > 2 and not re.match(r"^[\d.,\-/: ]+$", s):
                # Not a pure number/date/coordinate
                if s not in ("Du lịch ẩm thực", "Làng nghề", "Di tích lịch sử - văn hóa cấp quốc gia"):
                    name = s
                    name_col = i
                    break

        if not name or name in seen:
            continue

        # Find lat/lng: scan all cells for Vietnam-range values
        lat, lng = 0.0, 0.0
        for i, v in enumerate(cells):
            f = parse_coord(v)
            if 8.0 <= f <= 24.0 and lat == 0.0:
                lat = f
            elif 102.0 <= f <= 110.0 and lng == 0.0:
                lng = f
            if lat and lng:
                break

        if lat == 0.0 or lng == 0.0:
            continue

        city = nearest_city(lat, lng)
        if not city:
            continue

        # Find other fields heuristically
        loai = ""
        desc = ""
        district = ""
        hours = ""
        price = 0.0
        rating = 0.0

        for i, s in enumerate(str_cells):
            if i == name_col:
                continue
            if not s:
                continue
            if any(k in s.lower() for k in ["ẩm thực", "làng nghề", "di tích", "du lịch"]):
                loai = s
            elif len(s) > 30 and not desc:
                desc = s[:500]
            elif re.search(r"\d{1,2}:\d{2}", s) and not hours:
                hours = parse_hours(s)
            elif re.match(r"^[\d.,\-–/ ]+$", s) and "." in s and not price:
                p = parse_price(s)
                if p > 0:
                    price = p
            elif re.match(r"^\d+\.\d+$", s):
                try:
                    f = float(s)
                    if 1.0 <= f <= 5.0 and not rating:
                        rating = f
                except ValueError:
                    pass

        # Find district (typically a short string near city)
        for i, s in enumerate(str_cells):
            if s and s.lower() not in ("đà nẵng", "hội an", "huế") and len(s) < 30 and not re.match(r"[\d:+\-]", s) and i != name_col and not district:
                if any(k in s.lower() for k in ["hải châu", "sơn trà", "ngũ hành sơn", "cẩm lệ", "liên chiểu", "thanh khê", "điện", "hòa"]):
                    district = s
                    break

        seen.add(name)
        category = classify_category(loai) if loai else "food"
        if category not in ("food", "cafe"):
            category = "food"
        environment = classify_environment(loai, "")

        # Guess meal_time from operating hours
        meal_time = []
        if hours:
            open_h = 0
            m = re.match(r"(\d{1,2}):", hours)
            if m:
                open_h = int(m.group(1))
            if open_h <= 10:
                meal_time.append("breakfast")
            if open_h <= 14:
                meal_time.append("lunch")
            if not meal_time or open_h >= 11:
                meal_time.append("dinner")
        if not meal_time:
            meal_time = ["lunch", "dinner"]

        record = {
            "name": name,
            "category": "food",
            "type": "food",
            "environment": environment,
            "district": district or "",
            "city": city,
            "lat": round(lat, 6),
            "lng": round(lng, 6),
            "description": desc,
            "operating_hours": hours,
            "estimated_cost": price,
            "estimated_rating": round(rating, 1) if rating > 0 else 0,
            "car_parking": False,
            "motorbike_parking": True,
            "meal_time": meal_time,
            "season_tags": ["all_year"],
            "style": "heritage",
        }
        records.append(record)

    return records


def main():
    print(f"Reading {EXCEL_PATH.name}...")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    print("Processing locations sheet...")
    locations = read_locations_sheet(wb)

    print("Processing food sheet...")
    food = read_food_sheet(wb)

    # Stats
    loc_cities = {}
    for r in locations:
        loc_cities[r["city"]] = loc_cities.get(r["city"], 0) + 1
    food_cities = {}
    for r in food:
        food_cities[r["city"]] = food_cities.get(r["city"], 0) + 1

    print(f"\nLocations: {len(locations)} total")
    for city, count in sorted(loc_cities.items()):
        print(f"   {city.encode('ascii','replace').decode()}: {count}")

    print(f"\nFood: {len(food)} total")
    for city, count in sorted(food_cities.items()):
        print(f"   {city.encode('ascii','replace').decode()}: {count}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    loc_path = OUT_DIR / "locations.json"
    food_path = OUT_DIR / "food.json"

    with open(loc_path, "w", encoding="utf-8") as f:
        json.dump(locations, f, ensure_ascii=False, indent=2)
    print(f"Saved -> {loc_path} ({len(locations)} records)")

    with open(food_path, "w", encoding="utf-8") as f:
        json.dump(food, f, ensure_ascii=False, indent=2)
    print(f"Saved -> {food_path} ({len(food)} records)")
    print(f"Done! Total: {len(locations) + len(food)} records in seed files.")
    print("Next: docker compose build ai-engine && docker compose up -d ai-engine")


if __name__ == "__main__":
    main()
